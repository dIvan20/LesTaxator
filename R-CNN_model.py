import os
import sys
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from PIL import Image, ImageTk, ImageDraw, ImageFont
import torch
from torchvision.transforms import functional as F
from torchvision.models.detection import maskrcnn_resnet50_fpn
from torchvision.models.detection.faster_rcnn import FastRCNNPredictor
from torchvision.models.detection.mask_rcnn import MaskRCNNPredictor
import numpy as np
import cv2
import math
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


if getattr(sys, 'frozen', False):
    base_path = sys._MEIPASS
else:
    base_path = r"C:\Users\korot\Desktop\setup"


MODEL_WEIGHTS_PATH = os.path.join(base_path, "best_maskrcnn_model.pth")
SAVE_DIR = os.path.join(base_path, "results")
os.makedirs(SAVE_DIR, exist_ok=True)
DEVICE = torch.device('cuda' if torch.cuda.is_available() else 'cpu')


def y(x):
    a = 6.94441
    b = -6.17774
    c = 379.2784
    result = a * x ** 2 + b * x + c
    if result < 0:
        return float('nan')
    return result


def get_sbh_in_m2(x):
    sbh_cm2 = y(x)
    sbh_m2 = sbh_cm2 / 10000
    return sbh_m2


def calculate_V(Sbh_m2, c=1.27, a=0.246, c1=9.2, Kf=0.52):
    if Sbh_m2 <= 0:
        return float('nan')
    d_1_3 = 2 * math.sqrt(Sbh_m2 / math.pi)
    d_1_3_cm = d_1_3 * 100
    d_k = c * d_1_3_cm
    l_x = a * d_k + c1
    V_x = (math.pi * d_1_3 ** 2 / 4) * l_x * Kf
    return V_x


def load_model(num_classes=2):
    model = maskrcnn_resnet50_fpn(weights='DEFAULT')
    in_features = model.roi_heads.box_predictor.cls_score.in_features
    model.roi_heads.box_predictor = FastRCNNPredictor(in_features, num_classes)
    in_features_mask = model.roi_heads.mask_predictor.conv5_mask.in_channels
    hidden_layer = 256
    model.roi_heads.mask_predictor = MaskRCNNPredictor(in_features_mask, hidden_layer, num_classes)
    model.load_state_dict(torch.load(MODEL_WEIGHTS_PATH, map_location=DEVICE))
    model.to(DEVICE)
    model.eval()
    return model


def draw_mask_contour(draw, mask, outline_color=(255, 140, 0), width=3):
    mask_uint8 = (mask.astype(np.uint8)) * 255
    contours, _ = cv2.findContours(mask_uint8, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    for contour in contours:
        polygon = [tuple(point[0]) for point in contour]
        if len(polygon) > 1:
            polygon.append(polygon[0])
            draw.line(polygon, fill=outline_color, width=width)


class CrownSegmentationApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("ЛесТаксатор 1.0 (прототип)")
        self.state('zoomed')

        self.model = load_model()

        main_frame = tk.Frame(self)
        main_frame.pack(fill=tk.BOTH, expand=True)

        self.left_panel_width = 500
        self.left_frame = tk.Frame(main_frame, width=self.left_panel_width, bg="#f0f0f0")
        self.left_frame.pack(side=tk.LEFT, fill=tk.Y)
        self.left_frame.pack_propagate(False)

        label_font = ('Arial', 12)
        btn_font = ('Arial', 14)

        tk.Label(self.left_frame, text="Высота съемки (м):", font=label_font, bg="#f0f0f0").pack(pady=(20, 5))
        self.entry_height = tk.Entry(self.left_frame, width=15, font=label_font)
        self.entry_height.insert(0, "100")
        self.entry_height.pack(pady=(0, 20))

        tk.Label(self.left_frame, text="Выбор породы:", font=label_font, bg="#f0f0f0").pack(pady=(0, 5))
        tree_species = ["Авто", "Сосна сибирская", "Ель, пихта", "Береза"]
        self.combobox_species = ttk.Combobox(self.left_frame, values=tree_species, font=label_font, state="readonly")
        self.combobox_species.current(1)  # По умолчанию "Сосна сибирская"
        self.combobox_species.pack(pady=(0, 20), fill=tk.X, padx=10)

        tk.Label(self.left_frame, text="Координаты углов изображения (широта, долгота)", font=label_font, bg="#f0f0f0").pack(pady=(10, 5))

        self.coord_vars = []
        corners = ["Левый верхний", "Правый верхний", "Правый нижний", "Левый нижний"]
        for corner in corners:
            frame = tk.Frame(self.left_frame, bg="#f0f0f0")
            frame.pack(pady=2, padx=5, fill=tk.X)
            tk.Label(frame, text=corner + ":", font=label_font, width=16, anchor='w', bg="#f0f0f0").pack(side=tk.LEFT)
            lat_var = tk.StringVar()
            lon_var = tk.StringVar()
            self.coord_vars.append((lat_var, lon_var))
            tk.Entry(frame, textvariable=lat_var, width=12).pack(side=tk.LEFT, padx=(5, 2))
            tk.Entry(frame, textvariable=lon_var, width=12).pack(side=tk.LEFT, padx=(2, 5))

        btn_config = {'width': 20, 'height': 2, 'font': btn_font, 'padx': 10, 'pady': 5}
        self.btn_load = tk.Button(self.left_frame, text="Загрузить ортофотоплан", command=self.load_image, **btn_config)
        self.btn_load.pack(pady=(10, 5))

        self.btn_segment = tk.Button(self.left_frame, text="Выделить кроны", command=self.segment_crowns, state=tk.DISABLED, **btn_config)
        self.btn_segment.pack(pady=(0, 5))

        self.btn_edit = tk.Button(self.left_frame, text="Редактировать", command=self.edit_artifacts, state=tk.DISABLED, **btn_config)
        self.btn_edit.pack(pady=(0, 5))

        self.btn_save = tk.Button(self.left_frame, text="Сохранить результат", command=self.save_result, state=tk.DISABLED, **btn_config)
        self.btn_save.pack(pady=(0, 5))

        self.btn_excel = tk.Button(self.left_frame, text="Сохранить таблицу объёмов", command=self.save_table, state=tk.DISABLED, **btn_config)
        self.btn_excel.pack(pady=(0, 5))

        self.right_frame = tk.Frame(main_frame, bg="gray")
        self.right_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)

        self.canvas = tk.Canvas(self.right_frame, bg="gray")
        self.canvas.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        self.image = None
        self.img_display = None
        self.result_img = None
        self.crown_data = []
        self.current_scale = 1.0

    def load_image(self):
        file_path = filedialog.askopenfilename(
            title="Выберите изображение",
            filetypes=[("Image files", "*.jpg;*.jpeg;*.png;*.bmp;*.tif"), ("All files", "*.*")]
        )
        if not file_path:
            return
        try:
            self.image = Image.open(file_path).convert("RGB")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось открыть изображение:\n{e}")
            return
        self.display_image(self.image)
        self.btn_segment.config(state=tk.NORMAL)
        self.btn_save.config(state=tk.DISABLED)
        self.btn_excel.config(state=tk.DISABLED)
        self.btn_edit.config(state=tk.NORMAL)
        self.result_img = None
        self.crown_data = []

    def display_image(self, pil_img):
        canvas_w = self.right_frame.winfo_width()
        canvas_h = self.right_frame.winfo_height()
        if canvas_w < 10 or canvas_h < 10:
            canvas_w, canvas_h = 800, 600
        img_w, img_h = pil_img.size
        scale = min(canvas_w / img_w, canvas_h / img_h, 1.0)
        new_w, new_h = int(img_w * scale), int(img_h * scale)
        resized = pil_img.resize((new_w, new_h), Image.LANCZOS)
        self.img_display = ImageTk.PhotoImage(resized)
        self.canvas.config(width=new_w, height=new_h)
        self.canvas.delete("all")
        self.canvas.create_image(0, 0, anchor="nw", image=self.img_display)
        self.current_scale = scale

    def coords_to_latlon(self, x_rel, y_rel):
        try:
            lat0 = float(self.coord_vars[0][0].get())
            lon0 = float(self.coord_vars[0][1].get())
            lat1 = float(self.coord_vars[1][0].get())
            lon1 = float(self.coord_vars[1][1].get())
            lat2 = float(self.coord_vars[2][0].get())
            lon2 = float(self.coord_vars[2][1].get())
            lat3 = float(self.coord_vars[3][0].get())
            lon3 = float(self.coord_vars[3][1].get())
        except ValueError:
            return None, None
        lat_top = lat0 + x_rel * (lat1 - lat0)
        lat_bottom = lat3 + x_rel * (lat2 - lat3)
        lat = lat_top + y_rel * (lat_bottom - lat_top)
        lon_top = lon0 + x_rel * (lon1 - lon0)
        lon_bottom = lon3 + x_rel * (lon2 - lon3)
        lon = lon_top + y_rel * (lon_bottom - lon_top)
        return lat, lon

    def segment_crowns(self):
        if self.image is None:
            messagebox.showwarning("Внимание", "Сначала загрузите изображение!")
            return
        try:
            height_m = float(self.entry_height.get())
            corrected_height_m = height_m / 2
            if height_m <= 0:
                raise ValueError()
        except ValueError:
            messagebox.showerror("Ошибка", "Введите корректное положительное число для высоты съемки")
            return

        horizontal_coverage_m = height_m
        img_w, img_h = self.image.size
        pixel_size_m = corrected_height_m / img_w
        PIXEL_TO_M2 = pixel_size_m ** 2

        max_size = 1024
        scale = min(max_size / img_w, max_size / img_h, 1.0)
        if scale < 1.0:
            img_resized = self.image.resize((int(img_w * scale), int(img_h * scale)), Image.LANCZOS)
        else:
            img_resized = self.image

        img_tensor = F.to_tensor(img_resized).to(DEVICE)

        try:
            with torch.no_grad():
                outputs = self.model([img_tensor])
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка при обработке модели:\n{e}")
            return

        output = outputs[0]
        scores = output['scores'].cpu().numpy()
        masks = output['masks'].cpu().numpy()
        boxes = output['boxes'].cpu().numpy()

        threshold = 0.5
        keep = scores >= threshold
        masks = masks[keep]
        boxes = boxes[keep]

        pil_result = self.image.copy()
        draw = ImageDraw.Draw(pil_result)
        try:
            font = ImageFont.truetype("arial.ttf", 50)
        except:
            font = ImageFont.load_default()

        inv_scale = 1.0 / scale
        self.crown_data = []

        for i, (mask, box) in enumerate(zip(masks, boxes), start=1):
            mask_bin = mask[0] > 0.5
            mask_resized = cv2.resize(mask_bin.astype(np.uint8), (img_w, img_h), interpolation=cv2.INTER_NEAREST).astype(bool)

            contour_width = max(1, min(int(5 / self.current_scale), 5))
            draw_mask_contour(draw, mask_resized, width=contour_width)

            xmin, ymin, xmax, ymax = box * inv_scale
            xmin, ymin, xmax, ymax = map(int, (xmin, ymin, xmax, ymax))
            center_x = (xmin + xmax) / 2
            center_y = (ymin + ymax) / 2

            draw.text((center_x, center_y), str(i), fill=(255, 140, 0), font=font)

            x_rel = center_x / img_w
            y_rel = center_y / img_h
            lat, lon = self.coords_to_latlon(x_rel, y_rel)

            # Восстановление определения площади кроны через контур и масштаб
            mask_uint8 = (mask_resized.astype(np.uint8)) * 255
            contours, _ = cv2.findContours(mask_uint8, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
            if len(contours) == 0:
                continue

            contour = max(contours, key=cv2.contourArea)
            contour = contour.squeeze()
            if contour.ndim != 2 or contour.shape[0] < 3:
                continue

            x = contour[:, 0]
            y = contour[:, 1]
            area_px = 0.5 * np.abs(np.dot(x, np.roll(y, 1)) - np.dot(y, np.roll(x, 1)))

            # Перевод площади в м²
            area_m2 = area_px * (pixel_size_m ** 2)

            if area_m2 > 23:
                print(f"Крона {i} отвергнута из-за подозрительно большой площади: {area_m2:.2f} м²")
                continue

            sbh_m2 = get_sbh_in_m2(area_m2)
            sbh_m2 *=1.7
            V = calculate_V(sbh_m2) if sbh_m2 and not np.isnan(sbh_m2) else float('nan')

            print(f"Crown {i}: Площадь кроны (м²): {area_m2:.4f}, Сечение ствола (м²): {sbh_m2:.6f}, Объем (м³): {V if not math.isnan(V) else 'NaN'}, Широта: {lat}, Долгота: {lon}")

            self.crown_data.append({
                "Номер кроны": i,
                "Площадь кроны (м²)": round(area_m2, 4),
                "Объем древесины (м³)": round(V, 4) if not math.isnan(V) else None,
                "Широта": round(lat, 6) if lat is not None else None,
                "Долгота": round(lon, 6) if lon is not None else None,
            })

        self.result_img = pil_result
        self.display_image(pil_result)
        self.btn_save.config(state=tk.NORMAL)
        self.btn_excel.config(state=tk.NORMAL)

    def edit_artifacts(self):
        if self.result_img is None:
            messagebox.showwarning("Внимание", "Сначала выполните сегментацию!")
            return
        messagebox.showinfo("Редактирование", f"Удаление артефактов для породы: {self.combobox_species.get()} ещё не реализовано.")

    def save_result(self):
        if self.result_img is None:
            messagebox.showwarning("Внимание", "Нет результата для сохранения!")
            return
        save_path = filedialog.asksaveasfilename(
            initialdir=SAVE_DIR,
            title="Сохранить изображение с выделенными кронами",
            defaultextension=".jpg",
            filetypes=[("JPEG", "*.jpg"), ("PNG", "*.png")]
        )
        if not save_path:
            return
        try:
            self.result_img.save(save_path)
            messagebox.showinfo("Успех", f"Изображение сохранено:\n{save_path}")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось сохранить изображение:\n{e}")

    def save_table(self):
        if not self.crown_data:
            messagebox.showwarning("Внимание", "Нет результатов для вывода в таблицу!")
            return
        df = pd.DataFrame(self.crown_data)
        save_path = filedialog.asksaveasfilename(
            initialdir=SAVE_DIR,
            title="Сохранить таблицу объемов",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")]
        )
        if not save_path:
            return
        df.to_excel(save_path, index=False)

        try:
            workbook = load_workbook(save_path)
            sheet = workbook.active
            widths = {
                1: 20,
                2: 35,
                3: 35,
                4: 25,
                5: 25,
            }
            for col_num, width in widths.items():
                col_letter = get_column_letter(col_num)
                sheet.column_dimensions[col_letter].width = width
            workbook.save(save_path)
            messagebox.showinfo("Успех", f"Таблица сохранена с заданной шириной столбцов:\n{save_path}")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось установить ширину столбцов:\n{e}")


if __name__ == "__main__":
    app = CrownSegmentationApp()
    app.mainloop()
